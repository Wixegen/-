from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import os
import time
import openpyxl
import asyncio
import concurrent.futures


def Search_places(name_url): 
    name = star = review = adres = phoneNumber = hours = "Нет данных"
    service = Service(ChromeDriverManager().install())
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", False)
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    driver = webdriver.Chrome(service=service, options=options) 
    url = name_url
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "[class*='scroll _width_narrow']"))
        )
        time.sleep(2)
        
        inf = driver.find_elements(By.CSS_SELECTOR, "[class*='card-title-view__header-content']")
        if inf:
            name = inf[0].text.strip()
            print(f"Название: {name}")
        
        inf = driver.find_elements(By.CSS_SELECTOR, "[class*='business-rating-badge-view__rating-text']")
        if inf:
            star = inf[0].text.strip()
            print(f"Рейтинг: {star}")
        
        inf = driver.find_elements(By.CSS_SELECTOR, "[class*='business-header-rating-view__text _clickable']")
        if inf:
            review = inf[0].text.strip()
            print(f"Отзывы: {review}")
        
        inf = driver.find_elements(By.CSS_SELECTOR, "[class*='business-contacts-view__address']")
        if inf:
            adres = inf[0].text.strip()
            print(f"Адрес: {adres}")
        
        inf = driver.find_elements(By.CSS_SELECTOR, "[class*='card-phones-view__phone-number']")
        if inf:
            phoneNumber = inf[0].text.strip()
            phoneNumber = phoneNumber.replace("Показать телефон", "").strip()
            print(f"Телефон: {phoneNumber}")
        
        inf = driver.find_elements(By.CSS_SELECTOR, "[class*='business-working-status-view']")
        if inf:
            hours = inf[0].text.strip()
            print(f"Часы: {hours}")
        
        return (name, star, review, adres, phoneNumber, hours, url)
        
    except Exception as e:
        print(f"Ошибка: {e}")
        return ("Ошибка", "", "", "", "", "", url)
    finally:
        driver.quit()


def Excel(all_data, file_name="results.xlsx"):
    if os.path.exists(file_name):
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ данных с карт"
        headers = ["Название", "Оценка", "Отзывы", "Адрес", "Телефон", "Часы работы", "Ссылка"]
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header)
    
    next_row = ws.max_row + 1
    for data in all_data:
        for col, value in enumerate(data, 1):
            ws.cell(row=next_row, column=col, value=value)
        next_row += 1
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    wb.save(file_name)
    print(f"Сохранено в {file_name}")


async def main():
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    file_path = os.path.join(downloads, "parser_carts.txt")
    
    # Читаем ссылки
    if not os.path.exists(file_path):
        os.makedirs(downloads, exist_ok=True)
        with open(file_path, "w", encoding='utf-8') as f:
            f.write("Сюда пишем ссылки(Это сообщение удалить)")
        print(f"📝 Файл создан: {file_path}")
        print("   Вставьте ссылки и запустите скрипт снова")
        return  # ← ВАЖНО: выходим, если файла нет
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
        content = content.replace("Сюда пишем ссылки(Это сообщение удалить)", "")
        links = [line.strip() for line in content.split('\n') if line.strip()]
    
    if not links:
        print("📭 Нет ссылок для обработки")
        return
    
    print(f"Найдено {len(links)} ссылок")
    print("Запускаю асинхронный парсинг...")
    
    # Асинхронный запуск в пуле потоков
    loop = asyncio.get_event_loop()
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        # Создаём задачи
        tasks = [loop.run_in_executor(executor, Search_places, link) for link in links]
        
        # Ждём выполнения всех
        results = await asyncio.gather(*tasks)
    
    # Сохраняем результаты
    Excel(results)
    print(f"Готово, Обработано {len(results)} ссылок")


if __name__ == "__main__":
    asyncio.run(main())